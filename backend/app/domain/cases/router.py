"""
사건 관리 API
- 사건 목록/검색
- 실제 사건 엑셀 업로드 (국내+해외 시트 자동 인식)
- 고객 DB 엑셀 업로드 (연락처)
- 엑셀 템플릿 다운로드
"""
import io
from datetime import date, datetime

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.domain.cases.models import Case, Party

router = APIRouter()


# ----------------------------------------------------------------
# 사건 목록 + 검색
# ----------------------------------------------------------------
@router.get("")
async def list_cases(
    q: str | None = Query(default=None, description="사건번호/고객사/발명명칭 검색"),
    limit: int = Query(default=50, le=200),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(Case).order_by(Case.filed_at.desc().nullslast()).limit(limit)
    if q:
        like = f"%{q}%"
        stmt = select(Case).where(
            or_(
                Case.case_number.ilike(like),
                Case.client_name.ilike(like),
                Case.title_ko.ilike(like),
                Case.title_en.ilike(like),
                Case.app_number.ilike(like),
            )
        ).limit(limit)

    result = await db.execute(stmt)
    cases = result.scalars().all()
    return [_case_to_dict(c) for c in cases]


# ----------------------------------------------------------------
# 템플릿 다운로드 — /{case_id} 와일드카드보다 먼저 등록
# ----------------------------------------------------------------
@router.get("/template")
async def download_template():
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "사건목록"
    headers = ["OurRef", "국문명칭", "영문명칭", "의뢰인", "권리", "출원번호", "등록번호", "국가코드", "담당변리사", "부서", "현재상태", "사건마감일"]
    ws.append(headers)
    ws.append(["PM24001KR", "인공지능 기반 특허 분석 시스템", "AI-based Patent Analysis System", "삼성전자", "특허", "10-2024-0012345", "", "KR", "김동일", "특허1부", "출원중", "2025-06-30"])

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2C3E8C")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=cases_template.xlsx"},
    )


# ----------------------------------------------------------------
# 사건 단건 조회
# ----------------------------------------------------------------
@router.get("/{case_id}")
async def get_case(case_id: str, db: AsyncSession = Depends(get_db)):
    # case_id 또는 case_number로 조회
    result = await db.execute(
        select(Case).where(
            or_(Case.case_number == case_id)
        )
    )
    case = result.scalar_one_or_none()
    if not case:
        raise HTTPException(status_code=404, detail="사건을 찾을 수 없습니다.")

    # 관계자 조회
    parties_result = await db.execute(
        select(Party).where(Party.case_id == case.id)
    )
    parties = parties_result.scalars().all()

    d = _case_to_dict(case)
    d["parties"] = [
        {"name": p.name, "email": p.email, "role": p.role, "org_name": p.org_name}
        for p in parties
    ]
    return d


# ----------------------------------------------------------------
# 실제 사건 엑셀 업로드 (국내+해외 자동 처리)
# ----------------------------------------------------------------
@router.post("/upload")
async def upload_cases(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    import pandas as pd

    content = await file.read()
    fname = (file.filename or "").lower()
    xl = None
    last_err = None

    # 확장자 우선, 실패 시 반대 엔진으로 재시도
    engines = ["xlrd", "openpyxl"] if fname.endswith(".xls") else ["openpyxl", "xlrd"]
    for engine in engines:
        try:
            xl = pd.ExcelFile(io.BytesIO(content), engine=engine)
            break
        except Exception as e:
            last_err = e
            continue

    if xl is None:
        raise HTTPException(
            status_code=400,
            detail=(
                f"파일 읽기 실패: {last_err}\n"
                "해결 방법: Excel에서 파일을 열고 '다른 이름으로 저장' → .xlsx 형식으로 저장 후 다시 업로드하세요."
            ),
        )

    total_created, total_updated, total_errors = 0, 0, []

    for sheet_name in xl.sheet_names:
        df = xl.parse(sheet_name, dtype=str)
        df = df.where(pd.notna(df), None)

        # OurRef 컬럼이 없으면 이 시트는 건너뜀
        if "OurRef" not in df.columns:
            total_errors.append({
                "sheet": sheet_name, "row": 0, "ref": "-",
                "error": f"OurRef 컬럼 없음 (발견된 컬럼: {', '.join(df.columns[:10].tolist())})"
            })
            continue

        # 해외 시트 여부 판단
        is_overseas = "국가코드" in df.columns

        for idx, row in df.iterrows():
            our_ref = _str(row.get("OurRef"))
            if not our_ref or our_ref.startswith("(사용X)"):
                continue

            try:
                result = await db.execute(select(Case).where(Case.case_number == our_ref))
                case = result.scalar_one_or_none()

                country = _str(row.get("국가코드")) if is_overseas else "KR"
                client_name = _str(row.get("의뢰인")) or _str(row.get("출원인")) or "-"

                fields = {
                    "case_number": our_ref,
                    "your_ref": _str(row.get("YourRef")),
                    "title_ko": _str(row.get("국문명칭")),
                    "title_en": _str(row.get("영문명칭")),
                    "client_name": client_name,
                    "applicant": _str(row.get("출원인")),
                    "applicant_contact": _str(row.get("출원인담당자")) or _str(row.get("출원인담당")),
                    "country": country or "KR",
                    "division": _str(row.get("구분")),
                    "case_type": _str(row.get("권리")),
                    "app_category": _str(row.get("출원구분")),
                    "app_kind": _str(row.get("출원종류")),
                    "attorney": _str(row.get("담당변리사")),
                    "department": _str(row.get("부서")),
                    "status": _str(row.get("현재상태")),
                    "app_number": _str(row.get("출원번호")),
                    "reg_number": _str(row.get("등록번호")),
                    "intl_app_number": _str(row.get("국제출원번호")),
                    "ipc": _str(row.get("IPC분류")),
                    "notes": _str(row.get("비고")),
                    "deadline": _date(row.get("사건마감일")),
                    "app_deadline": _date(row.get("출원마감일")),
                    "reg_deadline": _date(row.get("등록마감일")),
                    "annual_deadline": _date(row.get("연차마감일")),
                    "filed_at": _date(row.get("출원일")),
                    "registered_at": _date(row.get("등록일")),
                    "received_at": _date(row.get("접수일")),
                }

                if case:
                    for k, v in fields.items():
                        if v is not None:
                            setattr(case, k, v)
                    total_updated += 1
                else:
                    case = Case(**{k: v for k, v in fields.items() if v is not None})
                    case.case_number = our_ref
                    case.client_name = client_name
                    case.country = country or "KR"
                    db.add(case)
                    await db.flush()
                    total_created += 1

                # 출원인담당자 → Party
                contact = _str(row.get("출원인담당자")) or _str(row.get("출원인담당"))
                if contact and case.id:
                    ex = await db.execute(
                        select(Party).where(Party.case_id == case.id, Party.name == contact)
                    )
                    if not ex.scalar_one_or_none():
                        db.add(Party(case_id=case.id, name=contact, role="client_contact"))

            except Exception as e:
                total_errors.append({"sheet": sheet_name, "row": idx + 2, "ref": our_ref, "error": str(e)})

    await db.commit()
    return {
        "status": "ok",
        "created": total_created,
        "updated": total_updated,
        "total": total_created + total_updated,
        "errors": total_errors[:20],
    }


# ----------------------------------------------------------------
# 고객 DB 업로드 (연락처)
# ----------------------------------------------------------------
@router.post("/upload-contacts")
async def upload_contacts(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    import pandas as pd

    content = await file.read()
    try:
        engine = "xlrd" if file.filename.endswith(".xls") else "openpyxl"
        df = pd.read_excel(io.BytesIO(content), dtype=str, engine=engine)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"파일 읽기 실패: {e}")

    df = df.where(pd.notna(df), None)
    created, updated, skipped = 0, 0, 0

    for _, row in df.iterrows():
        email = _str(row.get("E-mail"))
        name = _str(row.get("고객명")) or _str(row.get("고객명(영문)"))
        company = _str(row.get("회사명"))
        role_raw = _str(row.get("고객구분")) or "client"

        if not email and not name:
            skipped += 1
            continue

        role = "opponent_agent" if role_raw and "대리인" in role_raw else "client"

        existing = None
        if email:
            result = await db.execute(select(Party).where(Party.email == email.lower()))
            existing = result.scalar_one_or_none()

        if existing:
            existing.name = name or existing.name
            existing.org_name = company or existing.org_name
            updated += 1
        else:
            db.add(Party(
                name=name,
                email=email.lower() if email else None,
                role=role,
                org_name=company,
            ))
            created += 1

    await db.commit()
    return {"status": "ok", "created": created, "updated": updated, "skipped": skipped, "total": created + updated}


# ----------------------------------------------------------------
# 헬퍼
# ----------------------------------------------------------------
def _case_to_dict(c: Case) -> dict:
    return {
        "id": str(c.id),
        "case_number": c.case_number,
        "your_ref": c.your_ref,
        "title_ko": c.title_ko,
        "title_en": c.title_en,
        "client_name": c.client_name,
        "applicant": c.applicant,
        "applicant_contact": c.applicant_contact,
        "country": c.country,
        "division": c.division,
        "case_type": c.case_type,
        "attorney": c.attorney,
        "department": c.department,
        "status": c.status,
        "app_number": c.app_number,
        "reg_number": c.reg_number,
        "deadline": c.deadline.isoformat() if c.deadline else None,
        "filed_at": c.filed_at.isoformat() if c.filed_at else None,
        "registered_at": c.registered_at.isoformat() if c.registered_at else None,
        "notes": c.notes,
        "ipc": c.ipc,
    }


def _str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() not in ("nan", "none", "") else None


def _date(val) -> date | None:
    if not val:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", ""):
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    # fallback: pandas handles Excel serial dates and other edge cases
    try:
        import pandas as pd
        return pd.to_datetime(s).date()
    except Exception:
        return None
